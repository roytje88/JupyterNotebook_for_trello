#!/usr/bin/env python
# coding: utf-8

# In[1]:


from matplotlib import pyplot as plt
import json
import requests
import pprint as pprint
from datetime import datetime
from datetime import date
import ipywidgets as widgets
import openpyxl
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
import os
import xlsxwriter
import xlrd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


# In[ ]:





# In[ ]:


if not os.path.exists('.variables'):
    key = input("Enter your API-key: ")
    token = input("Enter your API-token: ")
    allboards = json.loads(json.dumps(requests.get("https://api.trello.com/1/members/me/boards?key="+key+"&token="+token).json()))
    print('These are your boards: ')
    for i in allboards:
        if i['closed'] == False:
            print('Name: ',i['name'],', ID:',i['id'])
    boardid = input("Enter your board-id: ")
    
    print('These are the lists on the chosen board.')
    alllists = json.loads(json.dumps(requests.get("https://api.trello.com/1/boards/"+boardid+"/lists?key="+key+"&token="+token).json()))
    for i in alllists:
        if i['closed'] == False:
            print(i['name'])
    print()
    startlists = input("Enter the lists with cards that are not started (comma separated): ")
    print()
    blockedlists = input("Enter the lists with cards that are blocked(comma separated): ")
    print()
    doinglists = input("Enter the lists with the doing cards (comma separated): ")
    print()
    donelists = input("Enter the lists with cards that are done (comma separated): ")
    print()
    print()
    filenamewithoutextension = input("Enter the name for the excel file (without .xlsx)")
    filenamewithextension = filenamewithoutextension+".xlsx"
    print()
    email = input("Enter your e-mailaccount: ")
    print()
    password = input("Enter your e-mail password: ")
    print()
    toaddr = input("Enter e-mail of recipient: ")
    print()
    subj = input("Enter subject: ")    
    print()
    print()
    file = open('.variables', 'w')
    file = open('.variables', 'a')
    file.write(key)
    file.write('\n')
    file.write(token)
    file.write('\n')
    file.write(boardid)
    file.write('\n')
    file.write(startlists)
    file.write('\n')
    file.write(blockedlists)
    file.write('\n')
    file.write(doinglists)
    file.write('\n')
    file.write(donelists)
    file.write('\n')
    file.write(filenamewithextension)
    file.write('\n')
    file.write(email)
    file.write('\n')
    file.write(password)
    file.write('\n')
    file.write(toaddr)
    file.write('\n')
    file.write(subj)
    file.close()
    print('Your variables are saved in the file \'.variables\'. Please do not delete this file.')


# In[ ]:


f = open('.variables', 'r')
variables = f.read().splitlines()

lijstenbeginnen = []
lijstenblocked = []
lijstendoing = []
lijstendone = []




api_key = variables[0]
api_token = variables[1]
bordid = variables[2]
lijstenbeginnen = variables[3].split(',')
lijstenblocked = variables[4].split(',')
lijstendoing = variables[5].split(',')
lijstendone = variables[6].split(',')
excelfile = variables[7]
email = variables[8]
password = variables[9]
toaddr = variables[10]
subj = variables[11]

lijstenforscrum = []
for i in lijstenbeginnen:
    lijstenforscrum.append(i)
    
lijstenforscrum.extend(lijstenblocked)
lijstenforscrum.extend(lijstendoing)
lijstenforscrum.extend(lijstendone)


f.close()


# In[ ]:


keys = "key="+api_key+"&token="+api_token
trello_base_url = "https://api.trello.com/1/"
board_url = trello_base_url+"boards/"+bordid
url_cards = board_url+"/cards?attachments=true&customFieldItems=true&"+keys
url_lists = board_url+"/lists?"+keys
url_customfields = board_url+"/customFields?"+keys
url_labels = board_url+"/labels?"+keys
url_members = board_url+"/members?"+keys
statussen = ['Nog starten','Blocked','Doing','Done']


# In[ ]:


cards = json.loads(json.dumps(requests.get(url_cards).json()))
lists = json.loads(json.dumps(requests.get(url_lists).json()))
customfields = json.loads(json.dumps(requests.get(url_customfields).json()))
labels = json.loads(json.dumps(requests.get(url_labels).json()))
members = json.loads(json.dumps(requests.get(url_members).json()))


# In[3]:


def idtodate(cardid):
    hex = cardid[0:8]
    timestamp = int(hex,16)
    timedate = datetime.fromtimestamp(timestamp)
    return timedate


# In[2]:


def schrijfnaarexcel(countdict,excelfile,sheetname):
    if not os.path.exists(excelfile):
        workbook = xlsxwriter.Workbook(excelfile)
        worksheet = workbook.add_worksheet("Lijsten")
#        for i in statussen:
#            worksheet = workbook.add_worksheet(i)
        worksheet = workbook.add_worksheet("Vervallen")
        workbook.close() 
    counts = []
    index = []
    columns = []
    now = datetime.now()
    counts.append(now.strftime("%d-%m-%Y %H:%M"))
    columns.append('Date')
    for i,j in countdict.items():
        columns.append(i)
        counts.append(j)
    df_columnnames = pd.DataFrame([columns])
    df_counts = pd.DataFrame([counts],columns=columns)
    book = load_workbook(excelfile)
    writer = pd.ExcelWriter(excelfile,engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    try:
        valuecount = writer.sheets[sheetname]['B1'].value
    except:
        workbook = xlsxwriter.Workbook(excelfile)
        worksheet = workbook.add_worksheet(sheetname)
        valuecount = writer.sheets[sheetname]['B1'].value
    if valuecount == 0 or valuecount == None:
        df_columnnames.to_excel(writer,sheet_name=sheetname,startrow=1, index=False,header=False)
        writer.sheets[sheetname]['A1'].value = 'Counter'
        writer.sheets[sheetname]['B1'].value = 2
    valuecount = writer.sheets[sheetname]['B1'].value
    df_counts.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname]['B1'].value, index = False,header= False)
    writer.sheets[sheetname]['B1'].value = valuecount + 1
    writer.save()

